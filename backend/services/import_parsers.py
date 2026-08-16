from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from schemas.imports import ProductImportRecord, ProductSkuImport


@dataclass(frozen=True)
class ProductImportParseResult:
    records: list[ProductImportRecord]
    errors: list[dict[str, Any]]
    total_count: int


def parse_product_file(
    content: bytes,
    filename: str,
    *,
    default_inventory: int = 0,
) -> list[ProductImportRecord]:
    rows = _read_rows(content, filename)
    return [_normalize_record(row, default_inventory=default_inventory) for row in rows]


def parse_product_file_result(
    content: bytes,
    filename: str,
    *,
    default_inventory: int = 0,
) -> ProductImportParseResult:
    rows = _read_rows(content, filename)
    records: list[ProductImportRecord] = []
    errors: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, row in enumerate(rows, start=1):
        raw_external_id = str(row.get("external_id") or row.get("product_id") or "").strip()
        try:
            record = _normalize_record(row, default_inventory=default_inventory)
            if record.external_id in seen:
                raise ValueError("duplicate external_id in import file")
            seen.add(record.external_id)
            records.append(record)
        except Exception as exc:
            errors.append(
                {"row": index, "external_id": raw_external_id or None, "message": str(exc)}
            )
    return ProductImportParseResult(records=records, errors=errors, total_count=len(rows))


def _read_rows(content: bytes, filename: str) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".json":
        payload = json.loads(content.decode("utf-8-sig"))
        rows = payload if isinstance(payload, list) else [payload]
    elif suffix == ".csv":
        rows = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        rows = [dict(zip(headers, row, strict=False)) for row in values[1:]]
    else:
        raise ValueError("supported formats: .json, .csv, .xlsx")
    return rows


def _normalize_record(row: dict[str, Any], *, default_inventory: int) -> ProductImportRecord:
    knowledge = _object(row.get("rag_knowledge"))
    description = str(
        row.get("description")
        or row.get("marketing_description")
        or knowledge.get("marketing_description")
        or ""
    )
    external_id = str(row.get("external_id") or row.get("product_id") or "").strip()
    category = str(row.get("category") or "").strip()
    title = str(row.get("title") or "").strip()
    if not external_id or not category or not title:
        raise ValueError("external_id/product_id, title and category are required")

    raw_skus = _array(row.get("skus"))
    if raw_skus:
        skus = [
            ProductSkuImport(
                external_id=str(value.get("external_id") or value.get("sku_id") or "").strip(),
                name=str(value.get("name") or _sku_name(_object(value.get("properties"))) or "默认规格"),
                attributes=_object(value.get("attributes") or value.get("properties")),
                price=Decimal(str(value.get("price", row.get("base_price", 0)))),
                inventory=int(value.get("inventory", value.get("stock", default_inventory))),
                active=_boolean(value.get("active"), True),
            )
            for value in raw_skus
        ]
    else:
        skus = [
            ProductSkuImport(
                external_id=str(row.get("sku_id") or f"{external_id}-default"),
                price=Decimal(str(row.get("price", row.get("base_price", 0)))),
                inventory=int(row.get("inventory", row.get("stock", default_inventory))),
            )
        ]

    image_values = _array(row.get("image_urls"))
    image_path = row.get("image_path") or row.get("image_url")
    if image_path and not image_values:
        image_values = [image_path]

    return ProductImportRecord(
        external_id=external_id,
        title=title,
        brand=str(row.get("brand") or ""),
        category=category,
        sub_category=str(row.get("sub_category") or "") or None,
        description=description,
        attributes=_object(row.get("attributes")),
        image_urls=[str(value) for value in image_values if value],
        skus=skus,
        active=_boolean(row.get("active"), True),
    )


def _object(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {}
    return {}


def _array(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if isinstance(value, str) and value.strip():
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else []
    return []


def _boolean(value: Any, default: bool) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是"}


def _sku_name(properties: dict[str, Any]) -> str:
    return " / ".join(f"{key}:{value}" for key, value in properties.items())
